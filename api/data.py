"""
api/data.py  –  Vercel Serverless Function (Python runtime)
Reads CopaCadetePref_Dashboard.xlsx from /data/ and returns all data as JSON.
Deploy the Excel inside /data/ at the project root.
"""
import json, os, re
from http.server import BaseHTTPRequestHandler

try:
    import openpyxl
except ImportError:
    openpyxl = None


def parse_excel(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Goleadores ───────────────────────────────────────────
    ws = wb['🥅 Goleadores']
    goleadores = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None or not str(row[0]).strip().isdigit():
            continue
        pos, jugador, equipo, pj, goles, pen, media, *_ = list(row) + [None] * 8
        if jugador and goles is not None:
            goleadores.append({
                'pos':      int(pos),
                'jugador':  str(jugador).strip(),
                'equipo':   str(equipo).strip() if equipo else '',
                'pj':       int(pj)    if pj    else 0,
                'goles':    int(goles),
                'penaltis': int(pen)   if pen   else 0,
                'media':    round(float(media), 2) if media else 0,
            })

    # ── Resultados 16avos ────────────────────────────────────
    ws2 = wb['⚽ Resultados 16avos']
    resultados = []
    for row in ws2.iter_rows(min_row=3, values_only=True):
        cols = list(row) + [None] * 12
        fase, _, local, gl, _sep, gv, visitante, fecha, hora, campo, cesped, arbitro = cols[:12]
        if local and gl is not None:
            resultados.append({
                'fase':      str(fase).strip()      if fase      else '',
                'local':     str(local).strip(),
                'gl':        int(gl),
                'gv':        int(gv)                if gv is not None else 0,
                'visitante': str(visitante).strip() if visitante else '',
                'fecha':     str(fecha).strip()     if fecha     else '',
                'campo':     str(campo).strip()     if campo     else '',
                'cesped':    str(cesped).strip()    if cesped    else '',
                'arbitro':   str(arbitro).strip()   if arbitro   else '',
            })

    # ── Stats Equipos ────────────────────────────────────────
    ws3 = wb['📈 Stats Equipos']
    stats = []
    for row in ws3.iter_rows(min_row=3, values_only=True):
        cols = list(row) + [None] * 11
        equipo, pj, pg, pe, pp, gf, gc, dg, clasif = cols[:9]
        if equipo and str(equipo).strip() not in ('TOTALES', ''):
            stats.append({
                'equipo':     str(equipo).strip(),
                'pj':         int(pj)  if pj  else 0,
                'pg':         int(pg)  if pg  else 0,
                'pe':         int(pe)  if pe  else 0,
                'pp':         int(pp)  if pp  else 0,
                'gf':         int(gf)  if gf  else 0,
                'gc':         int(gc)  if gc  else 0,
                'dg':         int(dg)  if dg  else 0,
                'clasificado': '✅' in str(clasif) if clasif else False,
            })

    # ── Octavos ──────────────────────────────────────────────
    ws4 = wb['🔜 Octavos de Final']
    octavos = []
    for row in ws4.iter_rows(min_row=3, values_only=True):
        cols = list(row) + [None] * 8
        partido, local, visitante, campo_ida, fecha_ida, fecha_vuelta, campo_vuelta, _ = cols[:8]
        if local:
            octavos.append({
                'id':           int(partido)            if partido      else 0,
                'local':        str(local).strip(),
                'visitante':    str(visitante).strip()  if visitante    else '',
                'campoIda':     str(campo_ida).strip()  if campo_ida    else '',
                'fechaIda':     str(fecha_ida).strip()  if fecha_ida    else '',
                'fechaVuelta':  str(fecha_vuelta).strip() if fecha_vuelta else '',
                'campoVuelta':  str(campo_vuelta).strip() if campo_vuelta else '',
            })

    return {
        'goleadores': goleadores,
        'resultados': resultados,
        'stats':      stats,
        'octavos':    octavos,
    }


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        # CORS headers – allow the Vercel frontend to call this
        self.send_response(200)
        self.send_header('Content-Type',  'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Cache-Control', 'no-cache')
        self.end_headers()

        if openpyxl is None:
            self.wfile.write(json.dumps({'error': 'openpyxl not installed'}).encode())
            return

        # Resolve path: api/ is one level deep, Excel is at root /data/
        base   = os.path.dirname(os.path.abspath(__file__))
        xlsx   = os.path.join(base, '..', 'data',
                              'CopaCadetePref_Dashboard.xlsx')

        if not os.path.exists(xlsx):
            self.wfile.write(
                json.dumps({'error': f'Excel not found at {xlsx}'}).encode()
            )
            return

        try:
            data = parse_excel(xlsx)
            self.wfile.write(json.dumps(data, ensure_ascii=False).encode())
        except Exception as e:
            self.wfile.write(json.dumps({'error': str(e)}).encode())
